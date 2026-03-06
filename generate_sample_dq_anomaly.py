#!/usr/bin/env python3
"""
Generate a sample DQ_Anomaly.xlsx with 3 pre-filled rows (simulating existing BAF system data).
Place this file + matching PDFs into the input folder, then run scheduler.py to update it via IDP.

Demonstrates 3 IDP update scenarios:
  1. Field is EMPTY in Excel          → IDP fills value + source + confidence
  2. Field has a DIFFERENT value      → IDP overwrites value + source + confidence
  3. Field has the CORRECT value      → IDP leaves it unchanged

Usage:
    python generate_sample_dq_anomaly.py [--output-path PATH]
"""

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.excel_service import EXCEL_COLUMNS

OUTPUT_DEFAULT = Path("input/DQ_Anomaly.xlsx")

# ---------------------------------------------------------------------------
# Sample rows — simulating what already exists in the BAF system.
# Keys must match column names in EXCEL_COLUMNS.
#
# Policy numbers match the actual PDF filenames in the test set:
#   903-256-158 → Bryan McCartney
#   903-256-151 → Carmen Jones-Smith
#   903-256-33  → Louise Lightbourn
# ---------------------------------------------------------------------------
SAMPLE_ROWS = [
    {
        # ── Row 1: Bryan McCartney ──────────────────────────────────────────
        # Scenario: name + policy pre-filled (correct); address wrong;
        #           phones, DOB, NI_NUMBER empty.
        "PERNO":         "BAF-001",
        "FIRST_NAME":    "BRYAN",          # correct — IDP will leave unchanged
        "SURNAME":       "MCCARTNEY",      # correct — IDP will leave unchanged
        "POLICY_NO":     "903-256-158",    # correct — IDP will leave unchanged
        "POL_STAT":      "IF",             # In Force (system value, not in docs)
        "FREQ":          "Monthly",        # correct
        "ADDR1":         "140 MT ROYAL AVENUE",  # WRONG → IDP will overwrite with "140 MT. ROYAL AVE."
        # MIDDLE_NAME, DOB, NI_NUMBER, ADDR3, HOME_PHONES, MOBILE_PHONES intentionally empty
    },
    {
        # ── Row 2: Carmen Jones-Smith ───────────────────────────────────────
        # Scenario: name + policy pre-filled (correct); DOB wrong format;
        #           address and phones empty.
        "PERNO":         "BAF-002",
        "FIRST_NAME":    "CARMEN",         # correct — IDP will leave unchanged
        "SURNAME":       "JONES-SMITH",    # correct — IDP will leave unchanged
        "POLICY_NO":     "903-256-151",    # correct — IDP will leave unchanged
        "POL_STAT":      "IF",
        "DOB":           "17/03/1957",     # WRONG format → IDP will overwrite with "17 MAR 1957"
        # NI_NUMBER, ADDR1, ADDR3, HOME_PHONES, MOBILE_PHONES intentionally empty
    },
    {
        # ── Row 3: Louise Lightbourn ────────────────────────────────────────
        # Scenario: only name + policy pre-filled; most fields empty.
        "PERNO":         "BAF-003",
        "FIRST_NAME":    "LOUISE",         # correct — IDP will leave unchanged
        "SURNAME":       "LIGHTBOURN",     # correct — IDP will leave unchanged
        "POLICY_NO":     "903-256-33",     # correct — IDP will leave unchanged
        "POL_STAT":      "IF",
        # MIDDLE_NAME, DOB, NI_NUMBER, ADDR1, ADDR2 intentionally empty
    },
]


def _add_headers(ws, headers: list[str]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1D4ED8")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font


def generate(output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    _add_headers(ws, EXCEL_COLUMNS)

    for row_data in SAMPLE_ROWS:
        row = [row_data.get(col, "") or "" for col in EXCEL_COLUMNS]
        ws.append(row)

    # Autosize columns
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Sample DQ_Anomaly.xlsx written to: {output_path.resolve()}")
    print()
    print("Next steps:")
    print("  1. Copy the matching PDFs into the same folder:")
    print("       903-256-158_*.pdf  (Bryan McCartney)")
    print("       903-256-151_*.pdf  (Carmen Jones-Smith)")
    print("       903-256-33_*.pdf   (Louise Lightbourn)")
    print("  2. Run the scheduler:")
    print("       python scheduler.py --input-dir ./input --output-dir ./output")
    print("  3. Check output/DQ_Anomaly_V2.xlsx for the updated rows + IDP Status column.")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate a sample DQ_Anomaly.xlsx for the IDP update flow.")
    parser.add_argument(
        "--output-path",
        type=Path,
        default=OUTPUT_DEFAULT,
        help=f"Where to save the file (default: {OUTPUT_DEFAULT})",
    )
    args = parser.parse_args()
    generate(args.output_path)
    return 0


if __name__ == "__main__":
    sys.exit(main())
