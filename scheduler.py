#!/usr/bin/env python3
"""
Scheduler that monitors the input folder for new PDFs, processes them,
appends to a single Excel file (DQ_Anomaly_V2.xlsx), and deletes input files after processing.

Usage:
    python scheduler.py [--input-dir INPUT] [--output-dir OUTPUT] [--poll-interval SECS]

Environment:
    OPENAI_API_KEY  - Required for extraction
    OPENAI_MODEL    - Model to use (default: gpt-5.2)
"""

import argparse
import logging
import shutil
import sys
import time
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.config import INPUT_DIR, OUTPUT_DIR, LOG_FILE
from app.services.excel_service import (
    OUTPUT_EXCEL_NAME,
    append_rows_to_excel,
    update_excel_with_idp,
)
from app.services.extraction_service import extract_fields_from_images
from app.services.pdf_service import pdf_to_base64_images

logger = logging.getLogger(__name__)

# How often to check for files needing processing
DEFAULT_POLL_INTERVAL = 5


def process_pdf(pdf_path: Path) -> dict | None:
    """Process a single PDF and return row dict, or None on failure."""
    logger.info("=" * 60)
    logger.info("START  %s", pdf_path.name)
    logger.info("=" * 60)

    start = time.perf_counter()
    try:
        images = pdf_to_base64_images(pdf_path)
        if not images:
            logger.warning("No pages rendered from %s", pdf_path.name)
            return None

        extraction_result = extract_fields_from_images(images)
        elapsed = round(time.perf_counter() - start, 1)

        populated = sum(
            1 for v in extraction_result["extracted_fields"].values() if v is not None
        )
        total = len(extraction_result["extracted_fields"])
        logger.info(
            "DONE   %s | fields=%d/%d | time=%.1fs",
            pdf_path.name, populated, total, elapsed,
        )
        return {
            "extracted_fields": extraction_result["extracted_fields"],
            "field_source": extraction_result["field_source"],
            "field_confidence": extraction_result["field_confidence"],
            "time_taken": elapsed,
        }

    except Exception:
        elapsed = round(time.perf_counter() - start, 1)
        logger.exception("FAILED %s | time=%.1fs", pdf_path.name, elapsed)
        return None


def _find_input_excel(input_dir: Path) -> Path | None:
    """Return the first .xlsx file found in input_dir, or None."""
    xlsx_files = sorted(input_dir.glob("*.xlsx"))
    return xlsx_files[0] if xlsx_files else None


def _write_status_to_input(input_excel: Path, policy_statuses: dict) -> None:
    """Write IDP Status values back into the input Excel so the next poll skips them."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    wb = load_workbook(input_excel)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {name: idx + 1 for idx, name in enumerate(headers) if name}
    policy_no_col = header_map.get("POLICY_NO")
    if not policy_no_col:
        return

    if "IDP Status" not in header_map:
        new_col = ws.max_column + 1
        cell = ws.cell(row=1, column=new_col)
        cell.value = "IDP Status"
        cell.fill = PatternFill("solid", fgColor="1D4ED8")
        cell.font = Font(color="FFFFFF", bold=True)
        header_map["IDP Status"] = new_col

    status_col = header_map["IDP Status"]
    for row_num in range(2, ws.max_row + 1):
        policy_no = ws.cell(row=row_num, column=policy_no_col).value
        if not policy_no:
            continue
        policy_key = str(policy_no).strip().replace("/", "-")
        if policy_key in policy_statuses:
            ws.cell(row=row_num, column=status_col).value = policy_statuses[policy_key]

    wb.save(input_excel)
    logger.info("IDP Status written back to input: %s", input_excel.name)


def process_existing_excel(input_dir: Path, output_dir: Path, input_excel: Path) -> None:
    """
    Update flow: read an Excel file from input_dir, match each row's POLICY_NO
    to a PDF in the same folder, extract via IDP, update the row, save to output_dir.
    """
    from openpyxl import load_workbook
    wb = load_workbook(input_excel)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    header_map = {name: idx + 1 for idx, name in enumerate(headers) if name}
    policy_no_col = header_map.get("POLICY_NO")
    idp_status_col = header_map.get("IDP Status")

    if not policy_no_col:
        logger.error("No POLICY_NO column found in %s — skipping update flow", input_excel.name)
        return

    logger.info("=" * 60)
    logger.info("UPDATE FLOW  reading %s (%d data rows)", input_excel.name, ws.max_row - 1)
    logger.info("=" * 60)

    start = time.perf_counter()
    policy_extractions: dict = {}
    no_doc_policies: set[str] = set()
    pdfs_to_delete: list[Path] = []
    processed_pdfs: set[Path] = set()

    for row_num in range(2, ws.max_row + 1):
        policy_no = ws.cell(row=row_num, column=policy_no_col).value
        if not policy_no:
            continue

        data_row = row_num - 1  # human-readable row number (1-based, excluding header)

        # Normalise: Excel may use '/' while PDF filenames use '-'
        policy_key = str(policy_no).strip().replace("/", "-")

        # Skip rows already processed in a previous scheduler cycle
        if idp_status_col:
            existing_status = ws.cell(row=row_num, column=idp_status_col).value
            if existing_status:
                logger.info("ROW %d  POLICY_NO=%s  already processed (status=%s) — skipping",
                            data_row, policy_key, existing_status)
                continue

        logger.info("-" * 60)
        logger.info("ROW %d  Reading row from Excel...", data_row)
        logger.info("ROW %d  POLICY_NO = %s  Searching for matching PDF...", data_row, policy_key)

        # Already processed this policy in a previous row — reuse result
        if policy_key in policy_extractions or policy_key in no_doc_policies:
            logger.info("ROW %d  Already processed — reusing result", data_row)
            continue

        matching = sorted(input_dir.glob(f"*{policy_key}*.pdf"))
        if not matching:
            logger.warning("ROW %d  No document found for POLICY_NO=%s — marking as No Doc Found", data_row, policy_key)
            no_doc_policies.add(policy_key)
            continue

        pdf_path = matching[0]
        logger.info("ROW %d  Document matched: %s", data_row, pdf_path.name)
        logger.info("ROW %d  Sending document to IDP for extraction...", data_row)

        if pdf_path in processed_pdfs:
            continue

        extracted = process_pdf(pdf_path)
        processed_pdfs.add(pdf_path)

        if extracted is not None:
            policy_extractions[policy_key] = extracted
            pdfs_to_delete.append(pdf_path)
            logger.info("ROW %d  Extraction complete. Updating Excel row...", data_row)
        else:
            logger.warning("ROW %d  Extraction failed for %s", data_row, pdf_path.name)

    output_path = output_dir / OUTPUT_EXCEL_NAME
    if policy_extractions or no_doc_policies:
        policy_statuses = update_excel_with_idp(
            input_excel, output_path, policy_extractions, no_doc_policies
        )
        # Write IDP Status back to input Excel so next poll skips processed rows
        _write_status_to_input(input_excel, policy_statuses)
    else:
        logger.info("No unprocessed rows found — nothing to do")

    processed_dir = output_dir / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)
    for pdf_path in pdfs_to_delete:
        try:
            dest = processed_dir / pdf_path.name
            shutil.move(str(pdf_path), str(dest))
            logger.info("Moved to processed: %s", pdf_path.name)
        except OSError:
            logger.exception("Could not move %s", pdf_path)

    elapsed = time.perf_counter() - start
    logger.info("=" * 60)
    logger.info(
        "UPDATE SUMMARY: %d rows updated | %d no-doc rows | %.1fs",
        len(policy_extractions), len(no_doc_policies), elapsed,
    )
    logger.info("Output Excel : %s", output_path)
    logger.info("Processed PDFs: %s", processed_dir)
    logger.info("Log file     : %s", LOG_FILE)
    logger.info("=" * 60)


def process_input_folder(input_dir: Path, output_dir: Path) -> None:
    """
    Detect mode and process:
    - If any .xlsx is present in input_dir → update existing rows (update flow)
    - Otherwise → extract all PDFs and create/append to DQ_Anomaly_V2.xlsx (create flow)
    """
    input_excel = _find_input_excel(input_dir)
    if input_excel:
        process_existing_excel(input_dir, output_dir, input_excel)
        return

    pdf_files = sorted(input_dir.glob("*.pdf"))
    if not pdf_files:
        return

    logger.info("Found %d PDF(s) in %s", len(pdf_files), input_dir)

    start = time.perf_counter()
    rows: list[dict] = []
    for pdf_path in pdf_files:
        extracted = process_pdf(pdf_path)
        if extracted is not None:
            rows.append(extracted)
            try:
                pdf_path.unlink(missing_ok=True)
                logger.info("Deleted input: %s", pdf_path.name)
            except OSError:
                logger.exception("Could not delete %s", pdf_path)

    if rows:
        output_path = output_dir / OUTPUT_EXCEL_NAME
        append_rows_to_excel(output_path, rows)

    elapsed = time.perf_counter() - start
    logger.info("=" * 60)
    logger.info("SUMMARY: %d/%d succeeded in %.1fs", len(rows), len(pdf_files), elapsed)
    logger.info("Output Excel : %s", output_dir / OUTPUT_EXCEL_NAME)
    logger.info("Log file     : %s", LOG_FILE)
    logger.info("=" * 60)


def run_polling_loop(input_dir: Path, output_dir: Path, interval: int) -> None:
    """Poll input folder periodically and process any PDFs found."""
    logger.info(
        "Scheduler started. Watching %s every %ds. Output: %s",
        input_dir,
        interval,
        output_dir,
    )

    while True:
        try:
            process_input_folder(input_dir, output_dir)
        except Exception:
            logger.exception("Error in scheduler loop")

        time.sleep(interval)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Monitor input folder for PDFs, append to DQ_Anomaly_V2.xlsx, delete input after processing."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=INPUT_DIR,
        help=f"Input folder to watch (default: {INPUT_DIR})",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help=f"Output folder for Excel (default: {OUTPUT_DIR})",
    )
    parser.add_argument(
        "--poll-interval",
        type=int,
        default=DEFAULT_POLL_INTERVAL,
        help=f"Seconds between folder checks (default: {DEFAULT_POLL_INTERVAL})",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="Process once and exit instead of polling continuously",
    )
    args = parser.parse_args()

    input_dir = args.input_dir.resolve()
    output_dir = args.output_dir.resolve()

    if not input_dir.exists():
        logger.error("Input directory does not exist: %s", input_dir)
        return 1

    output_dir.mkdir(parents=True, exist_ok=True)

    process_input_folder(input_dir, output_dir)

    if args.once:
        return 0

    run_polling_loop(input_dir, output_dir, args.poll_interval)
    return 0  # unreachable


if __name__ == "__main__":
    sys.exit(main())
